"""Excel (.xlsx) report export for large tabular reports."""
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="0F172A")
META_FONT = Font(size=10, color="64748B")


def _auto_width(ws, col_count: int, row_count: int):
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        max_len = 12
        for row in range(1, min(row_count + 1, 120)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 48))
        ws.column_dimensions[letter].width = max_len + 2


def build_report_xlsx(
    *,
    title: str,
    description: str,
    period_from: str,
    period_to: str,
    generated_by: str,
    role: str,
    summary: Dict[str, Any],
    insights: List[str],
    rows: List[dict],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Period: {period_from} to {period_to}"
    ws["A2"].font = META_FONT
    ws["A3"] = f"Prepared by: {generated_by} ({role.replace('_', ' ').title()})"
    ws["A3"].font = META_FONT
    ws["A4"] = description
    ws["A4"].font = META_FONT

    row_idx = 6
    if summary:
        ws.cell(row=row_idx, column=1, value="Summary").font = Font(bold=True, size=11)
        row_idx += 1
        for key, val in summary.items():
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=val)
            row_idx += 1
        row_idx += 1

    if insights:
        ws.cell(row=row_idx, column=1, value="Insights").font = Font(bold=True, size=11)
        row_idx += 1
        for line in insights:
            ws.cell(row=row_idx, column=1, value=f"• {line}")
            row_idx += 1
        row_idx += 1

    data_start = row_idx
    if not rows:
        ws.cell(row=data_start, column=1, value="No records for the selected scope.")
    else:
        keys = list(rows[0].keys())
        for c, key in enumerate(keys, 1):
            cell = ws.cell(row=data_start, column=c, value=key)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r_i, row in enumerate(rows, data_start + 1):
            for c, key in enumerate(keys, 1):
                ws.cell(row=r_i, column=c, value=row.get(key, ""))
        _auto_width(ws, len(keys), data_start + len(rows))

    ws.freeze_panes = ws.cell(row=data_start + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
